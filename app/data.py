from pathlib import Path
from download import Download
LINKURL = "https://www.city.kofu.yamanashi.jp/joho/opendata/shisetsu/index.html"
FILEURL = [
    "https://www.city.kofu.yamanashi.jp/joho/opendata/shisetsu/documents/kosodatespot_20200401.csv",
]

import pandas as pd
from openpyxl import load_workbook
import neologdn
from datetime import datetime, timedelta
import re
from jeraconv import jeraconv
j2w = jeraconv.J2W()

def format_date(st):
    if isinstance(st, int):
        if st > 59:
            st -= 1
        return pd.Timestamp("1899-12-31") + pd.Timedelta(st, unit="d")
    elif isinstance(st[0], str):
        _q = re.search("(平成元年)(\d+月\d+日)",st)
        nen = j2w.convert(_q[1])
        gappi = _q[2]
        return datetime.strptime(str(nen) + "年" + gappi, "%Y年%m月%d日")
    else:
        return datetime.strptime(st, "%Y年%m月%d日")

class ReadDatas():
    BASE_DIR = Path(__file__).absolute().parent.parent
    DATA_DIR = BASE_DIR / "data"

    def __init__(self):
        if not self.DATA_DIR.exists():
            self.DATA_DIR.mkdir()
        
        self.createdate = datetime.now()
        
        self.fnames = list()
        for url in FILEURL:
            d = Download(url, self.DATA_DIR)
            d.download()
            self.fnames.append(self.DATA_DIR / d.name)
            # format = self.suggest_format(d.name)
    
    def suggest_format(self, fname):
        suf = Path(fname).suffix.replace(".", "")
        if suf == "xlsx":
            pass
        elif suf == "xls":
            pass
        else:
            pass
        return suf

    def from_excel(self):
        pass
   
    def create_dataframe(self, fname, header=0, colrange="A:A"):
        data = list()
        wb = load_workbook(fname)
        for sname in wb.sheetnames:
            _df = pd.read_excel(fname, sheet_name=sname, header=header,usecols=colrange)
            _df.columns = [neologdn.normalize(i) for i in _df.columns]
            # カラムの表記ゆれ修正
            _df = _df.rename(columns={'事業所名': '事業所名称'})
            #_df = _df.applymap(neologdn.normalize)
            data.append(_df)
        df = pd.concat(data, ignore_index=True)
        return df

    def create_dataframe(self, fname, header=0):
        df = pd.read_csv(fname, encoding="sjis", index_col=header)
        df = df.fillna("-")
        df.columns = [neologdn.normalize(i) for i in df.columns]
        def replace_n(st):
            return str(st).replace("\n"," ")
        df = df.applymap(replace_n)
        df = df.applymap(neologdn.normalize)
        return df        

    def main(self):
        data = list()
        for fname in self.fnames:
            data.append(self.create_dataframe(fname, header=0))
        self.df = pd.concat(data, ignore_index=True)

    def get_version(self):
        da = self.createdate.strftime("%Y/%m/%d")
        return f"This data created in {da}"

    def query(self, keywords):
        return self.df.loc[self.df["施設名"].str.contains(keywords) | self.df["住所"].str.contains(keywords)]
        
if __name__ == "__main__":
    d = ReadDatas()
    d.main()
